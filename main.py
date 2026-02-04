"""
УЛУЧШЕННЫЙ ОСНОВНОЙ СКРИПТ ПАРСИНГА ДАННЫХ FSA
С ПОЛНЫМ ИЗВЛЕЧЕНИЕМ ОБЛАСТИ АККРЕДИТАЦИИ И НАЦИОНАЛЬНОЙ ЧАСТИ
"""
import asyncio
import aiohttp
import time
import json
import sys
import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import csv

# Импортируем настройки
try:
    from config import TOKEN, API_BASE, BASE_URL
except ImportError:
    print("ОШИБКА: Создайте файл config.py с токеном!")
    sys.exit(1)

# Настройки
LINKS_FILE = "all_links.txt"
OUTPUT_FILE = f"FSA_реестры_полные_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
OUTPUT_CSV = f"FSA_реестры_полные_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
STATS_FILE = "parsing_stats.json"

CONCURRENCY = 25
REQUEST_TIMEOUT = 60
RETRY_COUNT = 3
RETRY_DELAY = 3

# Заголовки
BASE_HEADERS = {
    "accept": "application/json, text/plain, */*",
    "accept-language": "ru-RU,ru;q=0.9",
    "authorization": TOKEN,
    "cache-control": "no-cache",
    "pragma": "no-cache",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "referer": "https://pub.fsa.gov.ru/ral",
    "origin": "https://pub.fsa.gov.ru"
}

# Словарь для преобразования названий столбцов в русские
COLUMN_TRANSLATIONS = {
    # Основные поля
    "source_url": "Ссылка на компанию",
    "company_id": "ID компании",
    "parsing_timestamp": "Время парсинга",
    "processing_time": "Время обработки (сек)",
    "error": "Ошибка",
    "company_error": "Ошибка компании",
    "declaration_error": "Ошибка декларации",
    "declaration_doc_id": "ID документа декларации",
    
    # Поля компании
    "company.id": "ID",
    "company.fullName": "Полное наименование",
    "company.shortName": "Сокращенное наименование",
    "company.regDate": "Дата внесения в реестр",
    "company.status.name": "Статус",
    "company.idType": "Тип аккредитованного лица",
    "company.idAccStandard": "ID стандарта",
    
    # Аккредитация
    "company.regNumbers": "Уникальный номер аккредитации",
    "accreditation_number": "Номер аккредитации",
    "accreditation_begin_date": "Дата начала аккредитации",
    "accreditation_end_date": "Дата окончания аккредитации",
    "accreditation_file_id": "ID файла области аккредитации",
    
    # Контакты компании
    "company_phone": "Номер телефона",
    "company_email": "Адрес электронной почты",
    "company_website": "Адрес сайта",
    "company_fax": "Факс",
    
    # Адрес компании
    "company_address_full": "Адрес места осуществления деятельности",
    "company_address_postal_code": "Индекс",
    "company_address_region": "Регион",
    "company_address_city": "Город",
    "company_address_street": "Улица",
    "company_address_house": "Дом",
    
    # Руководитель компании
    "head_full_name": "ФИО руководителя",
    "head_surname": "Фамилия руководителя",
    "head_name": "Имя руководителя",
    "head_patronymic": "Отчество руководителя",
    "head_post": "Должность руководителя",
    "head_phone": "Номер телефона руководителя",
    
    # Данные заявителя
    "applicant_type": "Заявитель: Тип заявителя",
    "applicant_legal_form": "Заявитель: Организационно-правовая форма",
    "applicant_full_name": "Заявитель: Полное наименование",
    "applicant_short_name": "Заявитель: Сокращенное наименование",
    "applicant_inn": "Заявитель: ИНН",
    "applicant_kpp": "Заявитель: КПП",
    "applicant_ogrn": "Заявитель: ОГРН",
    "applicant_is_government": "Заявитель: Государственное предприятие",
    "applicant_is_foreign": "Заявитель: Иностранная организация",
    "applicant_tax_authority": "Заявитель: Наименование налогового органа",
    "applicant_tax_reg_date": "Заявитель: Дата постановки на учет в налоговом органе",
    "applicant_head_full_name": "Заявитель: ФИО руководителя",
    "applicant_head_post": "Заявитель: Должность руководителя",
    "applicant_phone": "Заявитель: Номер телефона",
    "applicant_email": "Заявитель: Адрес электронной почты",
    "applicant_address": "Заявитель: Адрес места нахождения",
    "applicant_postal_code": "Заявитель: Почтовый индекс",
    
    # Национальная часть Единого реестра
    "np_included": "Включен в национальную часть Единого реестра",
    "np_decision_number": "НЧ ЕР: Номер решения о включении",
    "np_decision_date": "НЧ ЕР: Дата решения о включении",
    "np_service_number": "НЧ ЕР: Номер государственной услуги",
    "np_service_date": "НЧ ЕР: Дата государственной услуги",
    "np_last_decision_number": "НЧ ЕР: Номер последнего решения",
    "np_last_decision_date": "НЧ ЕР: Дата последнего решения",
    "np_tech_regulations": "НЧ ЕР: Технический регламент ЕАЭС",
    "np_tn_ved_codes": "НЧ ЕР: Коды ТН ВЭД ЕАЭС",
    "np_has_right_eaeu": "НЧ ЕР: Право на проведение оценки (Решения ЕЭК № 620)",
    
    # Область аккредитации
    "scope_type": "Область аккредитации: Тип",
    "scope_accredited_entity": "Область аккредитации: Наименование аккредитованного лица",
    "scope_address": "Область аккредитации: Адрес места осуществления деятельности",
    "scope_region": "Область аккредитации: Регион",
    "scope_business_type": "Область аккредитации: Вид деятельности",
    
    # Поля области аккредитации (из businessLineTypes.fields)
    "scope_note": "Примечание",
    "scope_tn_ved_code": "Код ТН ВЭД (ЕАЭС)",
    "scope_measurement_type": "Вид/метод измерений",
    "scope_measurement_range": "Показатели/диапазон",
    "scope_methodology": "Методика",
    "scope_test_object": "Наименование объекта испытаний",
    "scope_measurement_group": "Группа средств измерений",
    "scope_error": "Погрешность",
}

def log(msg):
    """Логирование"""
    ts = time.strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")

def translate_column_name(column_name):
    """Перевод названия столбца на русский"""
    return COLUMN_TRANSLATIONS.get(column_name, column_name)

def extract_company_id(url):
    """Извлечение ID компании из URL"""
    try:
        url = url.rstrip('/')
        if url.endswith('/current-aa'):
            url = url[:-len('/current-aa')]
        parts = url.split('/')
        return parts[-1] if parts else ""
    except:
        return ""

def extract_doc_id(company_json):
    """Извлечение ID документа аккредитации"""
    try:
        acc = company_json.get("accreditation")
        if isinstance(acc, dict):
            doc_id = acc.get("idAccredScopeFile")
            if doc_id:
                return str(doc_id)
        return None
    except:
        return None

def extract_contacts(contacts_array):
    """Извлекает контакты из массива"""
    contacts = {
        'phone': '',
        'email': '',
        'fax': '',
        'website': ''
    }
    
    if not contacts_array:
        return contacts
    
    for contact in contacts_array:
        contact_type = contact.get('idType')
        value = contact.get('value', '')
        
        if contact_type == 1:  # Телефон
            contacts['phone'] = value
        elif contact_type == 2:  # Факс
            contacts['fax'] = value
        elif contact_type == 3:  # Веб-сайт
            contacts['website'] = value
        elif contact_type == 4:  # Email
            contacts['email'] = value
    
    return contacts

def extract_address(addresses_array, address_type=3):
    """Извлекает адрес из массива адресов"""
    address_data = {
        'full': '',
        'postal_code': '',
        'region': '',
        'city': '',
        'street': '',
        'house': ''
    }
    
    if not addresses_array:
        return address_data
    
    # Ищем адрес нужного типа (3 = место осуществления деятельности)
    target_address = None
    for addr in addresses_array:
        if addr.get('idType') == address_type:
            target_address = addr
            break
    
    # Если не нашли нужный тип, берем первый
    if not target_address and addresses_array:
        target_address = addresses_array[0]
    
    if target_address:
        address_data['full'] = target_address.get('fullAddress', '')
        address_data['postal_code'] = target_address.get('postCode', '')
        
        # Пробуем извлечь регион, город и т.д. из uniqueAddress или fullAddress
        unique_addr = target_address.get('uniqueAddress', '')
        if unique_addr:
            address_data['region'] = unique_addr.split(',')[0] if ',' in unique_addr else ''
    
    return address_data

def extract_head_person(company_data):
    """Извлекает данные руководителя"""
    head_data = {}
    
    try:
        head_person = company_data.get('headPerson', {})
        if head_person:
            surname = head_person.get('surname', '')
            name = head_person.get('name', '')
            patronymic = head_person.get('patronymic', '')
            
            head_data['head_surname'] = surname
            head_data['head_name'] = name
            head_data['head_patronymic'] = patronymic
            head_data['head_full_name'] = f"{surname} {name} {patronymic}".strip()
            
            # Контакты руководителя
            head_contacts = extract_contacts(head_person.get('contacts', []))
            head_data['head_phone'] = head_contacts['phone']
        
        # Должность руководителя
        head_data['head_post'] = company_data.get('headPost', '')
    
    except Exception as e:
        head_data['head_error'] = str(e)
    
    return head_data

def extract_accreditation_data(company_data):
    """Извлекает данные об аккредитации"""
    accred_data = {}
    
    try:
        accreditation = company_data.get('accreditation', {})
        if accreditation:
            accred_data['accreditation_file_id'] = accreditation.get('idAccredScopeFile', '')
        
        # Номер аккредитации из regNumbers
        reg_numbers = company_data.get('regNumbers', [])
        if reg_numbers and isinstance(reg_numbers, list):
            for reg in reg_numbers:
                if reg.get('active'):
                    accred_data['accreditation_number'] = reg.get('regNumber', '')
                    accred_data['accreditation_begin_date'] = reg.get('beginDate', '')
                    accred_data['accreditation_end_date'] = reg.get('endDate', '')
                    break
    
    except Exception as e:
        accred_data['accreditation_error'] = str(e)
    
    return accred_data

def extract_applicant_data(company_data):
    """Извлекает детальные данные заявителя"""
    applicant_fields = {}
    
    try:
        if 'applicant' not in company_data:
            return applicant_fields
        
        applicant = company_data['applicant']
        
        # Основные данные
        type_map = {1: 'Физическое лицо', 2: 'Юридическое лицо', 3: 'ИП'}
        applicant_fields['applicant_type'] = type_map.get(applicant.get('idType'), '')
        applicant_fields['applicant_legal_form'] = applicant.get('nameLegalForm', '')
        applicant_fields['applicant_full_name'] = applicant.get('fullName', '')
        applicant_fields['applicant_short_name'] = applicant.get('shortName', '')
        applicant_fields['applicant_inn'] = applicant.get('inn', '')
        applicant_fields['applicant_kpp'] = applicant.get('kpp', '')
        applicant_fields['applicant_ogrn'] = applicant.get('ogrn', '')
        
        # Признаки
        applicant_fields['applicant_is_government'] = "Да" if applicant.get('isGovernmentCompany') else "Нет"
        applicant_fields['applicant_is_foreign'] = "Да" if applicant.get('isForeignOrganization') else "Нет"
        
        # Налоговый орган
        applicant_fields['applicant_tax_authority'] = applicant.get('taxAuthorityName', '')
        applicant_fields['applicant_tax_reg_date'] = applicant.get('taxAuthorityRegDate', '')
        
        # Руководитель заявителя
        applicant_fields['applicant_head_post'] = applicant.get('headPost', '')
        
        # Формируем ФИО руководителя из person
        person = applicant.get('person', {})
        if person:
            surname = person.get('surname', '')
            name = person.get('name', '')
            patronymic = person.get('patronymic', '')
            applicant_fields['applicant_head_full_name'] = f"{surname} {name} {patronymic}".strip()
        
        # Контакты заявителя
        contacts = extract_contacts(applicant.get('contacts', []))
        applicant_fields['applicant_phone'] = contacts['phone']
        applicant_fields['applicant_email'] = contacts['email']
        
        # Адрес заявителя
        addresses = applicant.get('addresses', [])
        if addresses:
            addr = addresses[0]
            applicant_fields['applicant_address'] = addr.get('fullAddress', '')
            applicant_fields['applicant_postal_code'] = addr.get('postCode', '')
    
    except Exception as e:
        applicant_fields['applicant_parse_error'] = str(e)
    
    return applicant_fields

def extract_national_part(company_data):
    """Извлекает данные из Национальной части Единого реестра"""
    np_fields = {}
    
    try:
        if 'actualInfoNationalPart' not in company_data:
            np_fields['np_included'] = "Нет"
            return np_fields
        
        np_fields['np_included'] = "Да"
        np_data = company_data['actualInfoNationalPart']
        
        # Основные данные
        np_fields['np_decision_number'] = np_data.get('decisionNumber', '')
        np_fields['np_decision_date'] = np_data.get('decisionDate', '')
        np_fields['np_service_number'] = np_data.get('serviceNumber', '')
        np_fields['np_service_date'] = np_data.get('serviceDate', '')
        
        # Последнее решение
        np_fields['np_last_decision_number'] = np_data.get('actualDecisionNumber', '')
        np_fields['np_last_decision_date'] = np_data.get('actualDecisionDate', '')
        
        # Технические регламенты и ТН ВЭД из accredScopeUnstructList
        tech_regs = set()
        tn_ved_codes = set()
        
        scope_list = np_data.get('accredScopeUnstructList', [])
        for scope_item in scope_list:
            scope_unstruct = scope_item.get('accredScopeUnstruct', {})
            
            # Технические регламенты
            regs_ts = scope_unstruct.get('regulationsTs', [])
            for reg in regs_ts:
                if reg:
                    tech_regs.add(reg)
            
            # Коды ТН ВЭД
            tn_ved = scope_unstruct.get('nationalPartTnVed', [])
            for code in tn_ved:
                if code:
                    tn_ved_codes.add(code)
        
        np_fields['np_tech_regulations'] = " ; ".join(sorted(tech_regs))
        np_fields['np_tn_ved_codes'] = " ; ".join(sorted(tn_ved_codes))
        
        # Право на проведение оценки
        np_fields['np_has_right_eaeu'] = "Да" if tech_regs or tn_ved_codes else "Нет"
        
    except Exception as e:
        np_fields['np_parse_error'] = str(e)
    
    return np_fields

def extract_accreditation_scope(declaration_data):
    """
    Извлекает детальные данные области аккредитации из declaration.root
    """
    scope_fields = {}
    
    try:
        if 'root' not in declaration_data or not declaration_data['root']:
            return scope_fields
        
        # declaration.root - это массив
        all_notes = []
        all_tn_ved = []
        all_measurement_types = []
        all_ranges = []
        all_methodologies = []
        all_test_objects = []
        all_measurement_groups = []
        all_errors = []
        
        for root_item in declaration_data['root']:
            # Основная информация
            scope_fields['scope_type'] = root_item.get('name', '')
            scope_fields['scope_accredited_entity'] = root_item.get('accreditedEntityName', '')
            
            # Адреса и места деятельности
            addresses = root_item.get('addresses', [])
            for address in addresses:
                scope_fields['scope_address'] = address.get('name', '')
                region_rf = address.get('regionRf', {})
                if region_rf:
                    scope_fields['scope_region'] = region_rf.get('name', '')
                
                # businessLineTypes - самое важное!
                business_lines = address.get('businessLineTypes', [])
                for business_line in business_lines:
                    scope_fields['scope_business_type'] = business_line.get('name', '')
                    
                    # fields - содержит все ключевые данные
                    fields = business_line.get('fields', [])
                    for field in fields:
                        field_name = field.get('name', '')
                        field_values = field.get('values', [])
                        
                        if not field_values:
                            continue
                        
                        # Объединяем значения
                        values_str = " | ".join(str(v) for v in field_values if v)
                        
                        # Распределяем по категориям
                        field_lower = field_name.lower()
                        
                        if 'примечан' in field_lower:
                            all_notes.append(values_str)
                        elif 'тн вэд' in field_lower or 'тн-вэд' in field_lower:
                            all_tn_ved.append(values_str)
                        elif 'вид' in field_lower and 'измер' in field_lower:
                            all_measurement_types.append(values_str)
                        elif 'диапазон' in field_lower or 'показател' in field_lower:
                            all_ranges.append(values_str)
                        elif 'методик' in field_lower:
                            all_methodologies.append(values_str)
                        elif 'объект' in field_lower and 'испыт' in field_lower:
                            all_test_objects.append(values_str)
                        elif 'группа' in field_lower and 'средств' in field_lower:
                            all_measurement_groups.append(values_str)
                        elif 'погрешност' in field_lower:
                            all_errors.append(values_str)
        
        # Объединяем все собранные данные
        if all_notes:
            scope_fields['scope_note'] = " ; ".join(all_notes)
        if all_tn_ved:
            scope_fields['scope_tn_ved_code'] = " ; ".join(all_tn_ved)
        if all_measurement_types:
            scope_fields['scope_measurement_type'] = " ; ".join(all_measurement_types)
        if all_ranges:
            scope_fields['scope_measurement_range'] = " ; ".join(all_ranges)
        if all_methodologies:
            scope_fields['scope_methodology'] = " ; ".join(all_methodologies)
        if all_test_objects:
            scope_fields['scope_test_object'] = " ; ".join(all_test_objects)
        if all_measurement_groups:
            scope_fields['scope_measurement_group'] = " ; ".join(all_measurement_groups)
        if all_errors:
            scope_fields['scope_error'] = " ; ".join(all_errors)
    
    except Exception as e:
        scope_fields['scope_parse_error'] = str(e)
    
    return scope_fields

async def fetch_with_retry(session, url, params=None):
    """Запрос с повторными попытками"""
    for attempt in range(RETRY_COUNT):
        try:
            async with session.get(url, params=params) as resp:
                if resp.status == 200:
                    return await resp.json()
                elif resp.status == 401:
                    return {"_error": "AUTH_ERROR", "_status": 401}
                elif resp.status == 404:
                    return {"_error": "NOT_FOUND", "_status": 404}
                elif resp.status == 429:
                    wait = (attempt + 1) * 5
                    log(f"Слишком много запросов. Жду {wait} сек...")
                    await asyncio.sleep(wait)
                    continue
                else:
                    log(f"Ошибка {resp.status} для {url}")
                    if attempt < RETRY_COUNT - 1:
                        await asyncio.sleep(RETRY_DELAY)
                        continue
                    return {"_error": f"HTTP_{resp.status}", "_status": resp.status}
        except asyncio.TimeoutError:
            log(f"Таймаут на попытке {attempt + 1}")
            if attempt < RETRY_COUNT - 1:
                await asyncio.sleep(RETRY_DELAY)
                continue
            return {"_error": "TIMEOUT"}
        except Exception as e:
            log(f"Ошибка на попытке {attempt + 1}: {str(e)}")
            if attempt < RETRY_COUNT - 1:
                await asyncio.sleep(RETRY_DELAY)
                continue
            return {"_error": str(e)}
    
    return {"_error": "MAX_RETRIES"}

async def process_company(session, url, idx, total, stats, semaphore):
    """Обработка одной компании"""
    async with semaphore:
        start_time = time.time()
        
        # Логируем прогресс
        if idx % 100 == 0:
            progress = (idx / total) * 100
            log(f"Прогресс: {idx}/{total:,} ({progress:.1f}%)")
        
        # Подготовка строки данных
        row = {
            "source_url": url,
            "parsing_timestamp": datetime.now().isoformat()
        }
        
        try:
            # Извлекаем ID компании
            company_id = extract_company_id(url)
            if not company_id:
                row["error"] = "Неверный URL"
                stats["errors"] += 1
                return row
            
            row["company_id"] = company_id
            
            # 1. Получаем данные компании
            company_url = f"{API_BASE}/ral/common/companies/{company_id}"
            company_data = await fetch_with_retry(session, company_url)
            
            if "_error" in company_data:
                if company_data.get("_status") == 401:
                    log("ОШИБКА: Токен недействителен!")
                    stats["auth_errors"] += 1
                row["company_error"] = f"API ошибка: {company_data['_error']}"
                stats["api_errors"] += 1
            else:
                # Базовые данные компании
                row['company.id'] = company_data.get('id', '')
                row['company.fullName'] = company_data.get('fullName', '')
                row['company.shortName'] = company_data.get('shortName', '')
                row['company.regDate'] = company_data.get('regDate', '')
                
                # Статус
                status = company_data.get('status', {})
                if status:
                    row['company.status.name'] = status.get('name', '')
                
                # Тип и стандарт
                row['company.idType'] = company_data.get('idType', '')
                row['company.idAccStandard'] = company_data.get('idAccStandard', '')
                
                # Номер аккредитации из regNumbers
                reg_numbers = company_data.get('regNumbers', [])
                if reg_numbers:
                    active_reg = None
                    for reg in reg_numbers:
                        if reg.get('active'):
                            active_reg = reg
                            break
                    if not active_reg and reg_numbers:
                        active_reg = reg_numbers[0]
                    
                    if active_reg:
                        row['company.regNumbers'] = active_reg.get('regNumber', '')
                
                # Контакты компании
                contacts = extract_contacts(company_data.get('contacts', []))
                row['company_phone'] = contacts['phone']
                row['company_email'] = contacts['email']
                row['company_website'] = contacts['website']
                row['company_fax'] = contacts['fax']
                
                # Адрес компании
                address = extract_address(company_data.get('addresses', []))
                row['company_address_full'] = address['full']
                row['company_address_postal_code'] = address['postal_code']
                row['company_address_region'] = address['region']
                
                # Руководитель
                head_data = extract_head_person(company_data)
                row.update(head_data)
                
                # Данные об аккредитации
                accred_data = extract_accreditation_data(company_data)
                row.update(accred_data)
                
                # Детальные данные заявителя
                applicant_data = extract_applicant_data(company_data)
                row.update(applicant_data)
                
                # Данные национальной части
                national_part_data = extract_national_part(company_data)
                row.update(national_part_data)
                
                stats["success_companies"] += 1
            
            # 2. Получаем декларацию (если есть)
            doc_id = extract_doc_id(company_data)
            if doc_id and "_error" not in company_data:
                decl_url = f"{API_BASE}/oa/accreditation/declaration/view/"
                params = {
                    "docId": doc_id,
                    "alType": 5,
                    "validate": "false"
                }
                
                decl_data = await fetch_with_retry(session, decl_url, params)
                
                if "_error" not in decl_data:
                    # ВАЖНО: Детальные данные области аккредитации
                    scope_data = extract_accreditation_scope(decl_data)
                    row.update(scope_data)
                    
                    row["declaration_doc_id"] = doc_id
                    stats["success_declarations"] += 1
                else:
                    row["declaration_error"] = decl_data["_error"]
            
        except Exception as e:
            row["error"] = str(e)
            stats["exceptions"] += 1
        
        # Время обработки
        row["processing_time"] = round(time.time() - start_time, 2)
        stats["total_time"] += row["processing_time"]
        
        return row

def create_excel_with_headers(columns, filename):
    """Создает Excel файл с заголовками"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Реестры FSA"
        
        # Сортируем колонки
        sorted_columns = sorted(columns)
        
        # Переводим заголовки на русский
        russian_headers = [translate_column_name(col) for col in sorted_columns]
        
        # Записываем заголовки
        ws.append(russian_headers)
        
        # Сохраняем файл только с заголовками
        wb.save(filename)
        
        log(f"Создан файл с заголовками: {filename}")
        return sorted_columns
        
    except Exception as e:
        log(f"Ошибка при создании Excel файла: {str(e)}")
        return None

def append_data_to_excel(rows, sorted_columns, filename):
    """Добавляет данные в существующий Excel файл"""
    try:
        # Загружаем существующий файл
        wb = load_workbook(filename)
        ws = wb.active
        
        # Добавляем данные
        for row in rows:
            data_row = []
            for col in sorted_columns:
                value = row.get(col, "")
                # Обработка специальных случаев
                if value is None:
                    value = ""
                elif isinstance(value, (dict, list)):
                    try:
                        value = json.dumps(value, ensure_ascii=False)
                    except:
                        value = str(value)
                elif isinstance(value, bool):
                    value = "Да" if value else "Нет"
                data_row.append(value)
            ws.append(data_row)
        
        # Сохраняем
        wb.save(filename)
        
    except Exception as e:
        log(f"Ошибка при добавлении данных в Excel: {str(e)}")
        raise

def save_to_csv(rows, columns, filename):
    """Сохраняет данные в CSV файл"""
    try:
        sorted_columns = sorted(columns)
        russian_headers = [translate_column_name(col) for col in sorted_columns]
        
        with open(filename, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f, delimiter=';', quoting=csv.QUOTE_MINIMAL)
            
            # Записываем заголовки
            writer.writerow(russian_headers)
            
            # Записываем данные
            for row in rows:
                data_row = []
                for col in sorted_columns:
                    value = row.get(col, "")
                    if value is None:
                        value = ""
                    elif isinstance(value, (dict, list)):
                        try:
                            value = json.dumps(value, ensure_ascii=False)
                        except:
                            value = str(value)
                    elif isinstance(value, bool):
                        value = "Да" if value else "Нет"
                    data_row.append(str(value))
                writer.writerow(data_row)
        
        log(f"Данные сохранены в CSV: {filename}")
        
    except Exception as e:
        log(f"Ошибка при сохранении CSV: {str(e)}")

def save_stats(stats, total_companies):
    """Сохраняет статистику парсинга"""
    stats["end_time"] = datetime.now().isoformat()
    if total_companies > 0:
        stats["success_rate"] = round((stats["success_companies"] / total_companies) * 100, 2)
        stats["avg_time_per_company"] = round(stats["total_time"] / total_companies, 2)
    else:
        stats["success_rate"] = 0
        stats["avg_time_per_company"] = 0
    
    with open(STATS_FILE, "w", encoding="utf-8") as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)
    
    log(f"Статистика сохранена: {STATS_FILE}")

def print_results(stats):
    """Выводит итоговые результаты"""
    log("\n" + "=" * 70)
    log("ИТОГИ ПАРСИНГА:")
    log("=" * 70)
    log(f"   Всего компаний: {stats['total_companies']:,}")
    log(f"   Успешно обработано: {stats['success_companies']:,} ({stats['success_rate']}%)")
    log(f"   Деклараций получено: {stats['success_declarations']:,}")
    log(f"   Ошибок API: {stats['api_errors']}")
    log(f"   Ошибок авторизации: {stats['auth_errors']}")
    log(f"   Исключений: {stats['exceptions']}")
    log(f"   Общее время: {stats['total_time']:.2f} сек")
    log(f"   Среднее время на компанию: {stats['avg_time_per_company']:.2f} сек")
    log("=" * 70)

async def main():
    """Основная функция"""
    log("Запуск УЛУЧШЕННОГО парсера FSA")
    log("=" * 70)
    
    # Проверяем файл со ссылками
    if not Path(LINKS_FILE).exists():
        log(f"ОШИБКА: Файл {LINKS_FILE} не найден!")
        log("Сначала запустите collect_links_improved.py для сбора ссылок")
        return
    
    # Читаем ссылки
    with open(LINKS_FILE, "r", encoding="utf-8") as f:
        links = [line.strip() for line in f if line.strip()]
    
    total = len(links)
    log(f"Найдено ссылок для обработки: {total:,}")
    log(f"Параллельных запросов: {CONCURRENCY}")
    
    if total == 0:
        log("Нет ссылок для обработки")
        return
    
    # Статистика
    stats = {
        "start_time": datetime.now().isoformat(),
        "total_companies": total,
        "success_companies": 0,
        "success_declarations": 0,
        "errors": 0,
        "api_errors": 0,
        "auth_errors": 0,
        "exceptions": 0,
        "total_time": 0
    }
    
    # Базовые колонки
    base_columns = set(COLUMN_TRANSLATIONS.keys())
    
    log("Создаю Excel файл с заголовками...")
    sorted_columns = create_excel_with_headers(base_columns, OUTPUT_FILE)
    if not sorted_columns:
        log("ОШИБКА: Не удалось создать Excel файл")
        return
    
    # Собираем все данные
    all_rows = []
    all_columns = set(base_columns)
    
    # Создаем сессию
    timeout = aiohttp.ClientTimeout(total=REQUEST_TIMEOUT)
    semaphore = asyncio.Semaphore(CONCURRENCY)
    
    async with aiohttp.ClientSession(headers=BASE_HEADERS, timeout=timeout) as session:
        # Создаем задачи для всех ссылок
        tasks = []
        for i, url in enumerate(links, 1):
            task = process_company(session, url, i, total, stats, semaphore)
            tasks.append(task)
        
        # Обрабатываем результаты по мере готовности
        completed = 0
        batch_rows = []
        batch_size = 100
        
        for future in asyncio.as_completed(tasks):
            try:
                row = await future
                batch_rows.append(row)
                
                # Обновляем набор колонок
                all_columns.update(row.keys())
                
                completed += 1
                
                # Периодическое сохранение
                if completed % batch_size == 0:
                    log(f"Обработано: {completed:,}/{total:,} ({completed/total*100:.1f}%)")
                    
                    try:
                        # Создаем новый файл с полным набором колонок, если изменились
                        current_columns = sorted(all_columns)
                        if set(current_columns) != set(sorted_columns):
                            log("Обнаружены новые колонки, обновляю заголовки...")
                            sorted_columns = create_excel_with_headers(all_columns, OUTPUT_FILE)
                            # Перезаписываем все данные
                            append_data_to_excel(all_rows, sorted_columns, OUTPUT_FILE)
                        else:
                            # Просто добавляем новые данные
                            append_data_to_excel(batch_rows, sorted_columns, OUTPUT_FILE)
                    except Exception as e:
                        log(f"Ошибка при промежуточном сохранении: {str(e)}")
                        save_to_csv(batch_rows, all_columns, f"backup_{completed}.csv")
                    
                    all_rows.extend(batch_rows)
                    batch_rows = []
                    
            except Exception as e:
                log(f"Ошибка в задаче: {str(e)}")
                stats["exceptions"] += 1
    
    # Добавляем оставшиеся данные
    if batch_rows:
        try:
            current_columns = sorted(all_columns)
            if set(current_columns) != set(sorted_columns):
                sorted_columns = create_excel_with_headers(all_columns, OUTPUT_FILE)
                append_data_to_excel(all_rows, sorted_columns, OUTPUT_FILE)
            else:
                append_data_to_excel(batch_rows, sorted_columns, OUTPUT_FILE)
        except Exception as e:
            log(f"Ошибка при финальном сохранении: {str(e)}")
            save_to_csv(batch_rows, all_columns, "backup_final.csv")
        
        all_rows.extend(batch_rows)
    
    # Сохраняем в CSV для надежности
    log("Создаю резервную копию в CSV формате...")
    save_to_csv(all_rows, all_columns, OUTPUT_CSV)
    
    # Сохраняем статистику
    save_stats(stats, total)
    
    # Итоги
    print_results(stats)
    log(f"\nФайл с данными (Excel): {OUTPUT_FILE}")
    log(f"Файл с данными (CSV): {OUTPUT_CSV}")
    log(f"Статистика: {STATS_FILE}")
    log("Парсинг завершен успешно!")

if __name__ == "__main__":
    # Исправляем кодировку для Windows
    if sys.platform == "win32":
        import io
        try:
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
            sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
        except:
            pass
    
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        log("\nПарсинг прерван пользователем")
    except Exception as e:
        log(f"Критическая ошибка: {str(e)}")
        import traceback
        traceback.print_exc()
