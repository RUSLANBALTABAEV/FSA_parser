"""
ОСНОВНОЙ ПАРСЕР ДАННЫХ FSA С ПРОКСИ
"""
import asyncio
import aiohttp
import time
import json
import sys
import os
import re
import random
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv

from config import (
    LINKS_FILE, EXCEL_FILE, CSV_FILE, JSON_FILE, STATS_FILE, LOG_FILE,
    HEADERS, MAX_CONCURRENT_REQUESTS, REQUEST_TIMEOUT, 
    RETRY_ATTEMPTS, SAVE_EVERY, OUTPUT_FORMATS,
    PROXY, PROXY_LIST, PROXY_MODE, VERIFY_SSL
)

def setup_logging():
    """Настраивает логирование"""
    import logging
    
    logger = logging.getLogger('fsa_parser')
    logger.setLevel(logging.INFO)
    
    formatter = logging.Formatter(
        '[%(asctime)s] %(levelname)s: %(message)s',
        datefmt='%H:%M:%S'
    )
    
    file_handler = logging.FileHandler(LOG_FILE, encoding='utf-8')
    file_handler.setFormatter(formatter)
    
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(formatter)
    
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger

logger = setup_logging()

def log(msg):
    """Упрощенный лог"""
    logger.info(msg)

class ProxyManager:
    """Менеджер для работы с прокси"""
    
    def __init__(self):
        self.proxy_index = 0
        self.proxy_errors = {}
        
    def get_proxy(self):
        """Получает текущий прокси"""
        if PROXY_MODE == "none":
            return None
        
        elif PROXY_MODE == "single":
            return PROXY if PROXY else None
        
        elif PROXY_MODE == "rotate" and PROXY_LIST:
            if len(PROXY_LIST) == 1:
                return PROXY_LIST[0]
            
            # Ротация прокси
            proxy = PROXY_LIST[self.proxy_index]
            self.proxy_index = (self.proxy_index + 1) % len(PROXY_LIST)
            
            # Пропускаем прокси с ошибками
            if proxy in self.proxy_errors:
                if time.time() - self.proxy_errors[proxy] < 300:  # 5 минут блокировки
                    return self.get_proxy()  # Пробуем следующий
            
            return proxy
        
        return None
    
    def mark_proxy_error(self, proxy):
        """Помечает прокси как ошибочный"""
        if proxy:
            self.proxy_errors[proxy] = time.time()
            log(f"Прокси отмечен как ошибочный: {proxy[:50]}...")
    
    def mark_proxy_success(self, proxy):
        """Помечает прокси как рабочий"""
        if proxy in self.proxy_errors:
            del self.proxy_errors[proxy]
            log(f"Прокси восстановлен: {proxy[:50]}...")

class FSAParser:
    """Класс для парсинга данных FSA с прокси"""
    
    def __init__(self):
        self.stats = {
            "start_time": datetime.now().isoformat(),
            "total_links": 0,
            "processed": 0,
            "successful": 0,
            "failed": 0,
            "errors": {},
            "proxy_used": PROXY_MODE,
            "total_time": 0
        }
        self.all_data = []
        self.current_batch = []
        self.proxy_manager = ProxyManager()
        
        # Структура данных для Excel
        self.excel_headers = [
            "ID компании",
            "Ссылка",
            "Название компании",
            "Статус",
            "ИНН",
            "ОГРН",
            "КПП",
            "Дата регистрации",
            "Телефоны",
            "Email",
            "Веб-сайт",
            "Адрес",
            "Регион",
            "Город",
            "Руководитель",
            "Должность руководителя",
            "Ключевые слова",
            "Даты",
            "Размер страницы (КБ)",
            "Статус HTTP",
            "Время обработки (сек)",
            "Дата парсинга",
            "Использованный прокси"
        ]
        
        log(f"Режим прокси: {PROXY_MODE}")
        if PROXY_MODE != "none":
            proxy = self.proxy_manager.get_proxy()
            log(f"Используемый прокси: {proxy[:50]}..." if proxy else "Прокси не настроен")
    
    def extract_company_id(self, url):
        """Извлекает ID компании из URL"""
        try:
            parts = url.rstrip('/').split('/')
            return parts[-2] if parts[-1] == 'current-aa' else parts[-1]
        except:
            return ""
    
    def create_session(self):
        """Создает сессию с настройками прокси"""
        import ssl
        import aiohttp
        
        # Создаем коннектор с настройками SSL
        ssl_context = None
        if not VERIFY_SSL:
            ssl_context = ssl.create_default_context()
            ssl_context.check_hostname = False
            ssl_context.verify_mode = ssl.CERT_NONE
        
        connector = aiohttp.TCPConnector(
            limit=MAX_CONCURRENT_REQUESTS,
            ssl=ssl_context if not VERIFY_SSL else True
        )
        
        timeout = aiohttp.ClientTimeout(total=REQUEST_TIMEOUT)
        
        return connector, timeout
    
    async def make_request(self, session, url, company_id):
        """Выполняет запрос через прокси"""
        proxy = self.proxy_manager.get_proxy()
        
        for attempt in range(RETRY_ATTEMPTS):
            try:
                # Подготавливаем параметры запроса
                request_params = {
                    'url': url,
                    'timeout': REQUEST_TIMEOUT,
                    'headers': HEADERS,
                    'ssl': not VERIFY_SSL  # Отключаем проверку SSL если нужно
                }
                
                # Добавляем прокси если есть
                if proxy and PROXY_MODE != "none":
                    request_params['proxy'] = proxy
                
                # Выполняем запрос
                async with session.get(**request_params) as response:
                    self.proxy_manager.mark_proxy_success(proxy)
                    return response
                
            except aiohttp.ClientProxyConnectionError as e:
                log(f"Ошибка подключения к прокси {proxy[:30]}...: {str(e)[:100]}")
                self.proxy_manager.mark_proxy_error(proxy)
                proxy = self.proxy_manager.get_proxy()  # Пробуем другой прокси
                
                if attempt < RETRY_ATTEMPTS - 1:
                    wait_time = 2 ** attempt
                    log(f"Жду {wait_time} сек перед повторной попыткой...")
                    await asyncio.sleep(wait_time)
                else:
                    raise
            
            except aiohttp.ClientError as e:
                log(f"Ошибка клиента для {company_id}: {str(e)[:100]}")
                if attempt < RETRY_ATTEMPTS - 1:
                    await asyncio.sleep(2 ** attempt)
                else:
                    raise
            
            except Exception as e:
                log(f"Общая ошибка для {company_id}: {str(e)[:100]}")
                if attempt < RETRY_ATTEMPTS - 1:
                    await asyncio.sleep(2 ** attempt)
                else:
                    raise
        
        return None
    
    def parse_html_content(self, html, url, company_id):
        """Парсит данные из HTML страницы"""
        data = {
            "company_id": company_id,
            "url": url,
            "company_name": "",
            "company_status": "",
            "inn": "",
            "ogrn": "",
            "kpp": "",
            "reg_date": "",
            "phones": "",
            "emails": "",
            "website": "",
            "address": "",
            "region": "",
            "city": "",
            "head_name": "",
            "head_position": "",
            "keywords": "",
            "dates": "",
            "page_size_kb": 0,
            "http_status": 0,
            "processing_time": 0,
            "parsing_date": datetime.now().isoformat(),
            "proxy_used": ""
        }
        
        try:
            soup = BeautifulSoup(html, 'html.parser')
            
            # 1. Название компании
            title = soup.find('title')
            if title:
                data["company_name"] = title.get_text(strip=True)
            
            # 2. Весь текст для поиска реквизитов
            all_text = soup.get_text()
            
            # 3. ИНН (10 или 12 цифр)
            inn_pattern = r'\b\d{10}\b|\b\d{12}\b'
            inn_matches = re.findall(inn_pattern, all_text)
            if inn_matches:
                data["inn"] = inn_matches[0]
            
            # 4. ОГРН (13 цифр)
            ogrn_pattern = r'\b\d{13}\b'
            ogrn_matches = re.findall(ogrn_pattern, all_text)
            if ogrn_matches:
                data["ogrn"] = ogrn_matches[0]
            
            # 5. КПП (9 цифр)
            kpp_pattern = r'\b\d{9}\b'
            kpp_matches = re.findall(kpp_pattern, all_text)
            if kpp_matches:
                data["kpp"] = kpp_matches[0]
            
            # 6. Телефоны
            phone_pattern = r'\+7\s?\(?\d{3}\)?\s?\d{3}[\s-]?\d{2}[\s-]?\d{2}|\b8\s?\(?\d{3}\)?\s?\d{3}[\s-]?\d{2}[\s-]?\d{2}'
            phones = re.findall(phone_pattern, all_text)
            if phones:
                data["phones"] = "; ".join(list(set(phones))[:3])
            
            # 7. Email
            email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
            emails = re.findall(email_pattern, all_text)
            if emails:
                data["emails"] = "; ".join(list(set(emails)))
            
            # 8. Веб-сайт
            website_pattern = r'https?://[^\s/$.?#].[^\s]*'
            websites = re.findall(website_pattern, all_text)
            if websites:
                data["website"] = websites[0]
            
            # 9. Адреса
            address_keywords = ['ул.', 'улица', 'проспект', 'пр.', 'дом', 'д.', 'г.', 'город', 'индекс']
            lines = all_text.split('\n')
            addresses = []
            for line in lines[:100]:
                line_lower = line.lower()
                if any(keyword in line_lower for keyword in address_keywords) and len(line.strip()) > 10:
                    addresses.append(line.strip())
            if addresses:
                data["address"] = addresses[0]
                
                # Пытаемся определить регион и город
                for addr in addresses:
                    if 'москв' in addr.lower():
                        data["region"] = "Москва"
                        data["city"] = "Москва"
                        break
                    elif 'санкт-петербург' in addr.lower() or 'спб' in addr.lower():
                        data["region"] = "Санкт-Петербург"
                        data["city"] = "Санкт-Петербург"
                        break
                    elif 'область' in addr.lower():
                        parts = addr.split(',')
                        for part in parts:
                            if 'область' in part.lower():
                                data["region"] = part.strip()
                                break
            
            # 10. Руководитель
            head_keywords = ['директор', 'руководитель', 'генеральный', 'президент', 'управляющий']
            for line in lines[:50]:
                line_lower = line.lower()
                if any(keyword in line_lower for keyword in head_keywords):
                    data["head_position"] = line.strip()
                    idx = lines.index(line)
                    for i in range(max(0, idx-3), min(len(lines), idx+4)):
                        if len(lines[i].strip()) > 5 and lines[i].strip() != line.strip():
                            words = lines[i].strip().split()
                            if 2 <= len(words) <= 4:
                                data["head_name"] = lines[i].strip()
                                break
                    break
            
            # 11. Статус компании
            status_keywords = {
                "действует": ["действует", "действующ", "активн", "валидн"],
                "приостановлена": ["приостановлен", "приост", "временн"],
                "прекращена": ["прекращен", "аннулир", "отозван", "лишен"]
            }
            
            text_lower = all_text.lower()
            for status, keywords in status_keywords.items():
                if any(keyword in text_lower for keyword in keywords):
                    data["company_status"] = status
                    break
            
            # 12. Даты
            date_pattern = r'\b\d{2}\.\d{2}\.\d{4}\b'
            dates = re.findall(date_pattern, all_text)
            if dates:
                data["dates"] = "; ".join(dates[:5])
                data["reg_date"] = dates[0]
            
            # 13. Ключевые слова FSA
            fsa_keywords = ['аккредитация', 'реестр', 'fsa', 'рф', 'сертификация', 'стандарт']
            found_keywords = []
            for keyword in fsa_keywords:
                if keyword.lower() in text_lower:
                    found_keywords.append(keyword)
            if found_keywords:
                data["keywords"] = "; ".join(found_keywords)
        
        except Exception as e:
            data["error"] = str(e)[:100]
        
        return data
    
    async def fetch_company_data(self, session, url, company_id, semaphore):
        """Получает данные компании"""
        async with semaphore:
            start_time = time.time()
            
            data = {
                "company_id": company_id,
                "url": url,
                "company_name": "",
                "company_status": "",
                "inn": "",
                "ogrn": "",
                "kpp": "",
                "reg_date": "",
                "phones": "",
                "emails": "",
                "website": "",
                "address": "",
                "region": "",
                "city": "",
                "head_name": "",
                "head_position": "",
                "keywords": "",
                "dates": "",
                "page_size_kb": 0,
                "http_status": 0,
                "processing_time": 0,
                "parsing_date": datetime.now().isoformat(),
                "proxy_used": ""
            }
            
            proxy_used = None
            
            try:
                for attempt in range(RETRY_ATTEMPTS):
                    try:
                        proxy_used = self.proxy_manager.get_proxy()
                        data["proxy_used"] = proxy_used[:50] + "..." if proxy_used else ""
                        
                        # Выполняем запрос через прокси
                        response = await self.make_request(session, url, company_id)
                        
                        if response:
                            data["http_status"] = response.status
                            
                            if response.status == 200:
                                html = await response.text()
                                data["page_size_kb"] = len(html) / 1024
                                
                                # Парсим данные
                                parsed_data = self.parse_html_content(html, url, company_id)
                                
                                # Объединяем данные
                                for key in parsed_data:
                                    if parsed_data[key]:
                                        data[key] = parsed_data[key]
                                
                                self.stats["successful"] += 1
                                log(f"Успешно: {company_id} (прокси: {proxy_used[:30]}...)")
                                break
                            
                            elif response.status == 404:
                                data["company_status"] = "Не найдено"
                                self.stats["failed"] += 1
                                self.stats["errors"].setdefault("404", 0)
                                self.stats["errors"]["404"] += 1
                                break
                            
                            else:
                                data["company_status"] = f"Ошибка {response.status}"
                                if attempt < RETRY_ATTEMPTS - 1:
                                    await asyncio.sleep(2 ** attempt)
                                    continue
                                else:
                                    self.stats["failed"] += 1
                                    self.stats["errors"].setdefault(str(response.status), 0)
                                    self.stats["errors"][str(response.status)] += 1
                                    break
                    
                    except asyncio.TimeoutError:
                        data["company_status"] = "Таймаут"
                        if attempt < RETRY_ATTEMPTS - 1:
                            await asyncio.sleep(2 ** attempt)
                            continue
                        else:
                            self.stats["failed"] += 1
                            self.stats["errors"].setdefault("timeout", 0)
                            self.stats["errors"]["timeout"] += 1
                    
                    except Exception as e:
                        data["company_status"] = f"Ошибка: {str(e)[:50]}"
                        if attempt < RETRY_ATTEMPTS - 1:
                            await asyncio.sleep(2 ** attempt)
                            continue
                        else:
                            self.stats["failed"] += 1
                            self.stats["errors"].setdefault("other", 0)
                            self.stats["errors"]["other"] += 1
            
            except Exception as e:
                data["company_status"] = f"Критическая ошибка: {str(e)[:50]}"
                self.stats["failed"] += 1
            
            # Время обработки
            data["processing_time"] = round(time.time() - start_time, 2)
            
            return data
    
    def init_excel_file(self):
        """Инициализирует Excel файл с заголовками"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Компании FSA"
            
            # Записываем заголовки
            ws.append(self.excel_headers)
            
            # Стили для заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col in range(1, len(self.excel_headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Настраиваем ширину столбцов
            column_widths = {
                "A": 15, "B": 60, "C": 40, "D": 15, "E": 15,
                "F": 20, "G": 15, "H": 15, "I": 25, "J": 30,
                "K": 30, "L": 40, "M": 20, "N": 20, "O": 25,
                "P": 25, "Q": 30, "R": 20, "S": 15, "T": 15,
                "U": 15, "V": 20, "W": 40
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            wb.save(EXCEL_FILE)
            log(f"Создан Excel файл: {EXCEL_FILE}")
            
            return True
        
        except Exception as e:
            log(f"Ошибка создания Excel файла: {str(e)}")
            return False
    
    def append_to_excel(self, data):
        """Добавляет данные в Excel файл"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            
            # Подготавливаем строку
            row = [
                data.get("company_id", ""),
                data.get("url", ""),
                data.get("company_name", ""),
                data.get("company_status", ""),
                data.get("inn", ""),
                data.get("ogrn", ""),
                data.get("kpp", ""),
                data.get("reg_date", ""),
                data.get("phones", ""),
                data.get("emails", ""),
                data.get("website", ""),
                data.get("address", ""),
                data.get("region", ""),
                data.get("city", ""),
                data.get("head_name", ""),
                data.get("head_position", ""),
                data.get("keywords", ""),
                data.get("dates", ""),
                data.get("page_size_kb", ""),
                data.get("http_status", ""),
                data.get("processing_time", ""),
                data.get("parsing_date", ""),
                data.get("proxy_used", "")
            ]
            
            ws.append(row)
            wb.save(EXCEL_FILE)
            
            return True
        
        except Exception as e:
            log(f"Ошибка записи в Excel: {str(e)}")
            return False
    
    def save_to_csv(self):
        """Сохраняет данные в CSV файл"""
        try:
            with open(CSV_FILE, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f, delimiter=';')
                
                # Заголовки
                writer.writerow(self.excel_headers)
                
                # Данные
                for data in self.all_data:
                    row = [
                        data.get("company_id", ""),
                        data.get("url", ""),
                        data.get("company_name", ""),
                        data.get("company_status", ""),
                        data.get("inn", ""),
                        data.get("ogrn", ""),
                        data.get("kpp", ""),
                        data.get("reg_date", ""),
                        data.get("phones", ""),
                        data.get("emails", ""),
                        data.get("website", ""),
                        data.get("address", ""),
                        data.get("region", ""),
                        data.get("city", ""),
                        data.get("head_name", ""),
                        data.get("head_position", ""),
                        data.get("keywords", ""),
                        data.get("dates", ""),
                        data.get("page_size_kb", ""),
                        data.get("http_status", ""),
                        data.get("processing_time", ""),
                        data.get("parsing_date", ""),
                        data.get("proxy_used", "")
                    ]
                    writer.writerow(row)
            
            log(f"Данные сохранены в CSV: {CSV_FILE}")
            return True
        
        except Exception as e:
            log(f"Ошибка сохранения CSV: {str(e)}")
            return False
    
    def save_to_json(self):
        """Сохраняет данные в JSON файл"""
        try:
            with open(JSON_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.all_data, f, ensure_ascii=False, indent=2, default=str)
            
            log(f"Данные сохранены в JSON: {JSON_FILE}")
            return True
        
        except Exception as e:
            log(f"Ошибка сохранения JSON: {str(e)}")
            return False
    
    def save_statistics(self):
        """Сохраняет статистику парсинга"""
        try:
            self.stats["end_time"] = datetime.now().isoformat()
            self.stats["total_time"] = round(time.time() - time.mktime(
                datetime.fromisoformat(self.stats["start_time"]).timetuple()
            ), 2)
            
            if self.stats["processed"] > 0:
                self.stats["success_rate"] = round((self.stats["successful"] / self.stats["processed"]) * 100, 2)
                self.stats["avg_time_per_company"] = round(self.stats["total_time"] / self.stats["processed"], 2)
            
            # Добавляем информацию о прокси
            self.stats["proxy_mode"] = PROXY_MODE
            if PROXY_MODE == "single":
                self.stats["proxy"] = PROXY[:50] + "..." if PROXY else None
            elif PROXY_MODE == "rotate":
                self.stats["proxy_count"] = len(PROXY_LIST)
                self.stats["proxies"] = [p[:50] + "..." for p in PROXY_LIST]
            
            with open(STATS_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.stats, f, ensure_ascii=False, indent=2, default=str)
            
            log(f"Статистика сохранена: {STATS_FILE}")
        
        except Exception as e:
            log(f"Ошибка сохранения статистики: {str(e)}")
    
    def print_summary(self):
        """Выводит итоговую статистику"""
        log("\n" + "=" * 70)
        log("ИТОГИ ПАРСИНГА С ПРОКСИ:")
        log("=" * 70)
        log(f"   Режим прокси: {PROXY_MODE}")
        
        if PROXY_MODE == "single" and PROXY:
            log(f"   Использован прокси: {PROXY[:50]}...")
        elif PROXY_MODE == "rotate":
            log(f"   Количество прокси: {len(PROXY_LIST)}")
        
        log(f"   Всего компаний: {self.stats['total_links']:,}")
        log(f"   Обработано: {self.stats['processed']:,}")
        log(f"   Успешно: {self.stats['successful']:,}")
        log(f"   С ошибками: {self.stats['failed']:,}")
        
        if 'success_rate' in self.stats:
            log(f"   Успешность: {self.stats['success_rate']}%")
        
        log(f"   Общее время: {self.stats.get('total_time', 0):.1f} сек")
        
        if self.stats["processed"] > 0:
            avg_time = self.stats.get('avg_time_per_company', 0)
            log(f"   Среднее время на компанию: {avg_time:.2f} сек")
        
        # Ошибки
        if self.stats["errors"]:
            log(f"\n   Ошибки:")
            for error_type, count in self.stats["errors"].items():
                log(f"     {error_type}: {count}")
        
        # Файлы
        log(f"\n   Созданные файлы:")
        if os.path.exists(EXCEL_FILE):
            log(f"     Excel: {EXCEL_FILE}")
        if os.path.exists(CSV_FILE):
            log(f"     CSV: {CSV_FILE}")
        if os.path.exists(JSON_FILE):
            log(f"     JSON: {JSON_FILE}")
        if os.path.exists(STATS_FILE):
            log(f"     Статистика: {STATS_FILE}")
        
        log("=" * 70)
    
    async def process_all_companies(self):
        """Обрабатывает все компании из файла со ссылками"""
        if not os.path.exists(LINKS_FILE):
            log(f"Файл {LINKS_FILE} не найден!")
            log("Сначала запустите generate_links.py")
            return
        
        with open(LINKS_FILE, 'r', encoding='utf-8') as f:
            urls = [line.strip() for line in f if line.strip()]
        
        self.stats["total_links"] = len(urls)
        log(f"Найдено ссылок для обработки: {len(urls):,}")
        
        if len(urls) == 0:
            log("Нет ссылок для обработки")
            return
        
        # Инициализируем Excel файл
        if not self.init_excel_file():
            log("Не удалось создать Excel файл")
            return
        
        # Создаем сессию
        connector, timeout = self.create_session()
        semaphore = asyncio.Semaphore(MAX_CONCURRENT_REQUESTS)
        
        async with aiohttp.ClientSession(headers=HEADERS, connector=connector, timeout=timeout) as session:
            # Создаем задачи для всех ссылок
            tasks = []
            company_map = {}
            
            for i, url in enumerate(urls, 1):
                company_id = self.extract_company_id(url)
                if company_id:
                    task = asyncio.create_task(
                        self.fetch_company_data(session, url, company_id, semaphore)
                    )
                    tasks.append(task)
                    company_map[task] = (i, company_id, url)
            
            # Обрабатываем результаты
            total_tasks = len(tasks)
            log(f"Начинаю обработку {total_tasks} компаний через прокси...")
            
            for i, task in enumerate(asyncio.as_completed(tasks), 1):
                try:
                    data = await task
                    idx, company_id, url = company_map[task]
                    
                    self.stats["processed"] += 1
                    self.all_data.append(data)
                    
                    # Сохраняем в Excel
                    self.append_to_excel(data)
                    
                    # Прогресс
                    if i % 50 == 0:
                        progress = (i / total_tasks) * 100
                        log(f"Обработано: {i:,}/{total_tasks:,} ({progress:.1f}%) | "
                            f"Успешно: {self.stats['successful']:,} | "
                            f"Ошибок: {self.stats['failed']:,}")
                    
                    # Периодическое сохранение
                    if i % SAVE_EVERY == 0:
                        log(f"Промежуточное сохранение... ({i} записей)")
                        if "csv" in OUTPUT_FORMATS:
                            self.save_to_csv()
                        if "json" in OUTPUT_FORMATS:
                            self.save_to_json()
                        self.save_statistics()
                
                except Exception as e:
                    log(f"Ошибка обработки задачи: {str(e)}")
                    self.stats["failed"] += 1
            
            # Финальное сохранение
            log("Финальное сохранение данных...")
            if "csv" in OUTPUT_FORMATS:
                self.save_to_csv()
            if "json" in OUTPUT_FORMATS:
                self.save_to_json()
            self.save_statistics()

async def main():
    """Основная функция парсера"""
    log("=" * 70)
    log("ПАРСЕР ДАННЫХ FSA С ПРОКСИ")
    log("=" * 70)
    log(f"Режим прокси: {PROXY_MODE}")
    log(f"Проверка SSL: {'ВКЛ' if VERIFY_SSL else 'ВЫКЛ'}")
    log(f"Файл со ссылками: {LINKS_FILE}")
    log(f"Выходной Excel файл: {EXCEL_FILE}")
    
    parser = FSAParser()
    
    try:
        await parser.process_all_companies()
        parser.print_summary()
        log("\n✓ Парсинг завершен успешно!")
        log(f"✓ Основной файл: {EXCEL_FILE}")
    
    except KeyboardInterrupt:
        log("\n✗ Парсинг прерван пользователем")
        parser.save_statistics()
        parser.print_summary()
    
    except Exception as e:
        log(f"\n✗ Критическая ошибка: {str(e)}")
        import traceback
        traceback.print_exc()
        parser.save_statistics()
        parser.print_summary()

if __name__ == "__main__":
    if sys.platform == "win32":
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    
    asyncio.run(main())
